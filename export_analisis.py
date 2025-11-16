import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import os

def get_validation_status(measurement, prev_measurement=None):
    """
    Menentukan status validasi (OK, WARNING, DANGER) berdasarkan data pengukuran
    """
    issues = []

    # Cek data lengkap
    has_complete_data = (
        measurement.get('tgl_ukur') is not None and
        measurement.get('umur_bulan') is not None and
        measurement.get('berat_kg') is not None and
        measurement.get('tinggi_cm') is not None and
        measurement.get('cara_ukur') is not None
    )

    if not has_complete_data:
        if measurement.get('berat_kg') is None and measurement.get('tinggi_cm') is None:
            issues.append("Data berat dan tinggi kosong")
        return "WARNING", issues

    # Cek status tinggi rasional
    status_tb_rasional = measurement.get('status_tb_rasional', '')
    if status_tb_rasional == 'DANGER':
        issues.append("Tinggi badan menurun")
        return "DANGER", issues

    # Cek status berat dan tinggi tidak ideal
    status_bb = measurement.get('status_bb', '')
    status_tb = measurement.get('status_tb', '')

    if status_bb not in ['NORMAL', 'TIDAK LENGKAP'] or status_tb not in ['NORMAL', 'TIDAK LENGKAP']:
        if status_bb not in ['NORMAL', 'TIDAK LENGKAP']:
            issues.append(f"Berat tidak ideal: {status_bb}")
        if status_tb not in ['NORMAL', 'TIDAK LENGKAP']:
            issues.append(f"Tinggi tidak ideal: {status_tb}")
        return "WARNING", issues

    return "OK", issues

def generate_keterangan(status, issues, measurement, child_data, prev_measurement=None):
    """
    Generate keterangan berdasarkan status dan issues yang ditemukan
    """
    if status == "OK":
        return ""

    keterangan_list = []

    for issue in issues:
        if "Data berat dan tinggi kosong" in issue:
            keterangan_list.append("Data berat dan tinggi kosong")
        elif "Tinggi badan menurun" in issue:
            if prev_measurement and measurement.get('tinggi_cm') and prev_measurement.get('tinggi_cm'):
                keterangan_list.append(f"Tinggi menurun: {prev_measurement['tinggi_cm']}cm → {measurement['tinggi_cm']}cm")
            else:
                keterangan_list.append("Tinggi badan menurun")
        elif "Berat tidak ideal" in issue:
            who_ref = measurement.get('rentang_bb_ideal')
            if who_ref and measurement.get('berat_kg'):
                keterangan_list.append(f"Berat tidak ideal: {measurement['berat_kg']}kg (ideal: {who_ref})")
            else:
                keterangan_list.append(issue)
        elif "Tinggi tidak ideal" in issue:
            who_ref = measurement.get('rentang_tb_ideal')
            if who_ref and measurement.get('tinggi_cm'):
                keterangan_list.append(f"Tinggi tidak ideal: {measurement['tinggi_cm']}cm (ideal: {who_ref})")
            else:
                keterangan_list.append(issue)
        else:
            keterangan_list.append(issue)

    return "; ".join(keterangan_list)

def export_to_excel_analisis(data, output_path=None):
    """
    Export data anak yang sudah dianalisis ke format Excel dengan analisis status
    """
    try:
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"analisis_pertumbuhan_anak_{timestamp}.xlsx"

        # Buat workbook baru
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Analisis Pertumbuhan Anak"

        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)

        ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Hijau muda
        warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Oranye
        danger_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Merah muda

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')

        # Header columns
        headers = [
            "No", "Tempat", "NIK", "Nama Anak", "Tanggal Lahir",
            "Bulan", "Tanggal Ukur", "Umur (bulan)", "Berat (kg)", "Tinggi (cm)",
            "Cara Ukur", "Status Berat", "Status Tinggi", "Validasi Input", "Keterangan"
        ]

        # Set header row
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border

        # Data rows
        row_idx = 2
        child_counter = 1

        for child in data.get('children', []):
            # Extract identity info
            no = child.get('no', child_counter)
            tempat = child.get('tempat', '')
            nik = child.get('nik', '')
            nama_anak = child.get('nama_anak', '')
            tanggal_lahir = child.get('tanggal_lahir', '')

            # Get measurements
            measurements = child.get('measurements', [])

            # Sort measurements by umur_bulan if available
            measurements_sorted = sorted(measurements, key=lambda x: x.get('umur_bulan', 0))

            prev_measurement = None

            for measurement in measurements_sorted:
                # Get status and issues
                status, issues = get_validation_status(measurement, prev_measurement)

                # Generate keterangan
                keterangan = generate_keterangan(status, issues, measurement, child, prev_measurement)

                # Prepare row data
                row_data = [
                    no, tempat, nik, nama_anak, tanggal_lahir,
                    measurement.get('periode', ''),
                    measurement.get('tgl_ukur', ''),
                    measurement.get('umur_bulan', ''),
                    measurement.get('berat_kg', ''),
                    measurement.get('tinggi_cm', ''),
                    measurement.get('cara_ukur', ''),
                    measurement.get('status_bb', ''),
                    measurement.get('status_tb', ''),
                    status,
                    keterangan
                ]

                # Write row data
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border

                    # Set alignment based on column type
                    if col_idx in [1, 8]:  # No, Umur - center
                        cell.alignment = center_alignment
                    elif col_idx in [2, 3, 4, 5, 6, 7, 11]:  # Text columns - left
                        cell.alignment = left_alignment
                    else:  # Numbers - center
                        cell.alignment = center_alignment

                # Apply color based on validation status (column 14 - Validasi Input)
                validation_cell = ws.cell(row=row_idx, column=14)
                if status == "OK":
                    validation_cell.fill = ok_fill
                elif status == "WARNING":
                    validation_cell.fill = warning_fill
                elif status == "DANGER":
                    validation_cell.fill = danger_fill

                # Also apply color to Keterangan column (column 15)
                if keterangan:
                    keterangan_cell = ws.cell(row=row_idx, column=15)
                    if status == "OK":
                        keterangan_cell.fill = ok_fill
                    elif status == "WARNING":
                        keterangan_cell.fill = warning_fill
                    elif status == "DANGER":
                        keterangan_cell.fill = danger_fill

                row_idx += 1
                prev_measurement = measurement

            child_counter += 1

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Max width 50
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save the file
        wb.save(output_path)
        return True, output_path

    except Exception as e:
        return False, f"Error exporting to Excel: {str(e)}"

def export_analisis_from_json(json_data, output_filename=None):
    """
    Export analisis dari data JSON yang sudah diproses
    """
    try:
        if output_filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"Analisis_Pertumbuhan_Anak_{timestamp}.xlsx"

        # Ensure output directory exists
        output_dir = "exports"
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, output_filename)

        success, result = export_to_excel_analisis(json_data, output_path)

        if success:
            return True, output_path
        else:
            return False, result

    except Exception as e:
        return False, f"Error in export process: {str(e)}"

# Fungsi untuk testing
if __name__ == "__main__":
    # Sample data for testing
    sample_data = {
        "children": [
            {
                "no": 1,
                "tempat": "Posyandu Melati",
                "nik": "1234567890123456",
                "nama_anak": "Andi Pratama",
                "tanggal_lahir": "2023-01-15",
                "jenis_kelamin": "L",
                "measurements": [
                    {
                        "periode": "Januari 2024",
                        "tgl_ukur": "2024-01-15",
                        "umur_bulan": 12,
                        "berat_kg": 9.5,
                        "tinggi_cm": 75.2,
                        "cara_ukur": "BERDIRI",
                        "status_bb": "NORMAL",
                        "status_tb": "NORMAL",
                        "status_tb_rasional": "NO_BASELINE",
                        "rentang_bb_ideal": "8.9-10.8",
                        "rentang_tb_ideal": "71.0-79.2"
                    },
                    {
                        "periode": "Februari 2024",
                        "tgl_ukur": "2024-02-15",
                        "umur_bulan": 13,
                        "berat_kg": None,
                        "tinggi_cm": 74.0,
                        "cara_ukur": "BERDIRI",
                        "status_bb": "TIDAK LENGKAP",
                        "status_tb": "PENDEK",
                        "status_tb_rasional": "DANGER",
                        "rentang_bb_ideal": "9.1-11.0",
                        "rentang_tb_ideal": "72.5-80.8"
                    }
                ]
            }
        ]
    }

    # Test export
    success, result = export_analisis_from_json(sample_data)
    if success:
        print(f"Export berhasil: {result}")
    else:
        print(f"Export gagal: {result}")