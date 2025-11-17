import pandas as pd
import json
import os
from datetime import datetime
import openpyxl

# Global variable to store WHO data
who_table = None

def extract_period_names_from_merged_cells(file_path):
    """
    Extract period names from merged cells in Excel file
    Returns list of period names in order
    """
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        period_names = []

        # Iterate through merged cells to find period names
        for merge_range in ws.merged_cells.ranges:
            min_col, min_row, max_col, max_row = merge_range.min_col, merge_range.min_row, merge_range.max_col, merge_range.max_row

            # Only process merged cells in row 1 (index 1 in openpyxl, which is row 1 in Excel)
            if min_row == 1 and max_row == 1 and min_col >= 6:  # Start from column 6 (after identity columns)
                cell_value = ws.cell(row=min_row, column=min_col).value
                if cell_value:
                    period_names.append(str(cell_value).strip())

        # Sort period names by column position
        period_names_with_col = []
        for merge_range in ws.merged_cells.ranges:
            min_col, min_row, max_col, max_row = merge_range.min_col, merge_range.min_row, merge_range.max_col, merge_range.max_row
            if min_row == 1 and max_row == 1 and min_col >= 6:
                cell_value = ws.cell(row=min_row, column=min_col).value
                if cell_value:
                    period_names_with_col.append((min_col, str(cell_value).strip()))

        period_names_with_col.sort(key=lambda x: x[0])
        period_names = [name for col, name in period_names_with_col]

        wb.close()
        return period_names

    except Exception as e:
        print(f"Error extracting period names from merged cells: {str(e)}")
        # Fallback to default period names
        return []

def load_who_table():
    """
    Load WHO growth reference table from CSV file
    """
    global who_table
    if who_table is not None:
        return who_table

    try:
        who_file_path = os.path.join(os.path.dirname(__file__), 'data master', 'Tabel_Pertumbuhan_Anak_0-2_Tahun.csv')
        if os.path.exists(who_file_path):
            who_table = pd.read_csv(who_file_path, sep=';')
            print(f"WHO table loaded successfully from {who_file_path}")
        else:
            print(f"Warning: WHO table not found at {who_file_path}")
            who_table = pd.DataFrame()
    except Exception as e:
        print(f"Error loading WHO table: {str(e)}")
        who_table = pd.DataFrame()

    return who_table

def get_who_reference(umur_bulan, jenis_kelamin):
    """
    Get WHO reference data for age and gender
    Returns: (min_bb, max_bb, min_tb, max_tb) or None if out of range
    """
    if who_table is None or who_table.empty:
        return None

    try:
        umur_bulan = int(umur_bulan)
        if umur_bulan < 0 or umur_bulan > 59:  # WHO table covers 0-59 months
            return None

        # Find row for this age
        age_row = who_table[who_table['Umur'] == umur_bulan]
        if age_row.empty:
            return None

        age_data = age_row.iloc[0]

        # Select columns based on gender - handle None case
        if not jenis_kelamin:
            return None

        jk_upper = str(jenis_kelamin).upper()
        if jk_upper == 'L':
            bb_col = 'BB Ideal (L)'
            tb_col = 'PB Ideal (L)'
        elif jk_upper == 'P':
            bb_col = 'BB Ideal (P)'
            tb_col = 'PB Ideal (P)'
        else:
            return None

        # Parse ranges like "5.3-8.8" to min and max values
        def parse_range(range_str):
            try:
                if '-' in str(range_str):
                    min_val, max_val = str(range_str).split('-')
                    return float(min_val.strip()), float(max_val.strip())
                else:
                    return None, None
            except:
                return None, None

        min_bb, max_bb = parse_range(age_data[bb_col])
        min_tb, max_tb = parse_range(age_data[tb_col])

        return {
            'min_bb': min_bb,
            'max_bb': max_bb,
            'min_tb': min_tb,
            'max_tb': max_tb,
            'rentang_bb_ideal': f"{min_bb}-{max_bb}" if min_bb and max_bb else None,
            'rentang_tb_ideal': f"{min_tb}-{max_tb}" if min_tb and max_tb else None
        }
    except Exception as e:
        print(f"Error getting WHO reference: {str(e)}")
        return None

def assess_nutritional_status(berat_kg, tinggi_cm, who_ref):
    """
    Assess nutritional status based on WHO reference
    Returns: (status_bb, status_tb)
    """
    if who_ref is None:
        return "TIDAK LENGKAP", "TIDAK LENGKAP"

    status_bb = "TIDAK LENGKAP"
    status_tb = "TIDAK LENGKAP"

    # Assess weight
    if berat_kg is not None and who_ref['min_bb'] is not None and who_ref['max_bb'] is not None:
        if berat_kg < who_ref['min_bb']:
            status_bb = "KURANG"
        elif berat_kg > who_ref['max_bb']:
            status_bb = "LEBIH"
        else:
            status_bb = "NORMAL"
    elif berat_kg is None:
        status_bb = "TIDAK LENGKAP"
    else:
        status_bb = "TIDAK LENGKAP"

    # Assess height
    if tinggi_cm is not None and who_ref['min_tb'] is not None and who_ref['max_tb'] is not None:
        if tinggi_cm < who_ref['min_tb']:
            status_tb = "PENDEK"
        elif tinggi_cm > who_ref['max_tb']:
            status_tb = "TINGGI"
        else:
            status_tb = "NORMAL"
    elif tinggi_cm is None:
        status_tb = "TIDAK LENGKAP"
    else:
        status_tb = "TIDAK LENGKAP"

    return status_bb, status_tb

def validate_height_rationality(measurements):
    """
    Validate height rationality between consecutive measurements
    Updates measurements with status_tb_rasional and catatan_tb_rasional
    """
    for i, measurement in enumerate(measurements):
        if i == 0:
            # First measurement has no baseline
            measurement['status_tb_rasional'] = "NO_BASELINE"
            measurement['catatan_tb_rasional'] = "Tidak ada data sebelumnya untuk memverifikasi rasionalitas tinggi badan."
            continue

        tb_sekarang = measurement.get('tinggi_cm')
        cara_ukur = measurement.get('cara_ukur')
        cara_sekarang = cara_ukur.upper() if cara_ukur else ''

        # Get previous measurement with height data
        prev_measurement = None
        for j in range(i-1, -1, -1):
            if measurements[j].get('tinggi_cm') is not None:
                prev_measurement = measurements[j]
                break

        if prev_measurement is None:
            measurement['status_tb_rasional'] = "NO_BASELINE"
            measurement['catatan_tb_rasional'] = "Tidak ada data sebelumnya untuk memverifikasi rasionalitas tinggi badan."
            continue

        tb_sebelumnya = prev_measurement.get('tinggi_cm')
        cara_ukur_sebelumnya = prev_measurement.get('cara_ukur')
        cara_sebelumnya = cara_ukur_sebelumnya.upper() if cara_ukur_sebelumnya else ''

        # Validate rationality
        if tb_sekarang is None:
            measurement['status_tb_rasional'] = "NO_BASELINE"
            measurement['catatan_tb_rasional'] = "Tidak ada data tinggi badan untuk periode ini."
            continue

        if tb_sekarang < tb_sebelumnya:
            selisih = tb_sebelumnya - tb_sekarang
            if cara_sekarang == cara_sebelumnya:
                measurement['status_tb_rasional'] = "DANGER"
                measurement['catatan_tb_rasional'] = "Tinggi badan menurun, tidak rasional."
            else:
                if selisih < 1:
                    measurement['status_tb_rasional'] = "AMBIGU_METHODOLOGY"
                    measurement['catatan_tb_rasional'] = "Penurunan kecil bisa karena beda metode ukur."
                else:
                    measurement['status_tb_rasional'] = "DANGER"
                    measurement['catatan_tb_rasional'] = "Penurunan besar, kemungkinan data salah."
        elif tb_sekarang == tb_sebelumnya:
            measurement['status_tb_rasional'] = "NORMAL"
            measurement['catatan_tb_rasional'] = "Tinggi badan stabil dibanding bulan sebelumnya."
        else:  # tb_sekarang > tb_sebelumnya
            measurement['status_tb_rasional'] = "NORMAL"
            measurement['catatan_tb_rasional'] = "Pertumbuhan tinggi badan normal."

def apply_assessment_rules(child_data):
    """
    Apply WHO assessment rules and height rationality validation to child measurements
    Now applies assessment to ALL measurements (complete AND incomplete)
    """
    try:
        # Apply height rationality validation first
        validate_height_rationality(child_data['measurements'])

        # Apply WHO assessment for ALL measurements (complete AND incomplete)
        for measurement in child_data['measurements']:
            umur_bulan = measurement.get('umur_bulan')
            jenis_kelamin = child_data.get('jenis_kelamin')
            berat_kg = measurement.get('berat_kg')
            tinggi_cm = measurement.get('tinggi_cm')

            # Get WHO reference for this age and gender (even if incomplete data)
            # Skip if jenis_kelamin is None or umur_bulan is None
            if jenis_kelamin and umur_bulan is not None:
                who_ref = get_who_reference(umur_bulan, jenis_kelamin)
            else:
                who_ref = None

            # Initialize assessment fields
            measurement['status_bb'] = "TIDAK LENGKAP"
            measurement['status_tb'] = "TIDAK LENGKAP"
            measurement['rentang_bb_ideal'] = None
            measurement['rentang_tb_ideal'] = None

            # Always set WHO reference ranges if available
            if who_ref:
                # Set WHO reference ranges
                measurement['rentang_bb_ideal'] = who_ref['rentang_bb_ideal']
                measurement['rentang_tb_ideal'] = who_ref['rentang_tb_ideal']

                # Always assess nutritional status (even if data incomplete)
                status_bb, status_tb = assess_nutritional_status(berat_kg, tinggi_cm, who_ref)
                measurement['status_bb'] = status_bb
                measurement['status_tb'] = status_tb
            else:
                # No WHO reference - set appropriate status
                if jenis_kelamin is None:
                    measurement['status_bb'] = "TIDAK LENGKAP"
                    measurement['status_tb'] = "TIDAK LENGKAP"
                elif umur_bulan is None:
                    measurement['status_bb'] = "TIDAK LENGKAP"
                    measurement['status_tb'] = "TIDAK LENGKAP"
                else:
                    measurement['status_bb'] = "OUT_OF_RANGE"
                    measurement['status_tb'] = "OUT_OF_RANGE"

            # No need to change status for incomplete data - "TIDAK LENGKAP" is already appropriate
            # This ensures assessment data is displayed consistently for all measurements

    except Exception as e:
        print(f"Error applying assessment rules: {str(e)}")

def process_excel_to_json(file_path):
    """
    Convert Excel file to JSON format for Balita Growth data
    Supports multiple formats:
    1. PRD Format: Header di baris 1, sub-header di baris 2, data di baris 3+
    2. Header Format: Header TGL UKUR, UMUR, dll di baris 1, data di baris 2+
    3. Direct Data: Data langsung tanpa header
    """
    try:
        # Load WHO reference table first
        load_who_table()

        # Detect format
        format_type, format_description = detect_excel_format(file_path)

        if format_type == 'prd_format':
            return process_prd_format(file_path)
        elif format_type == 'header_format':
            return process_header_format(file_path)
        elif format_type == 'direct_data':
            return process_direct_data_format(file_path)
        else:
            return {
                'error': f'Format tidak didukung: {format_description}',
                'file_name': os.path.basename(file_path),
                'format_detected': format_type
            }

    except Exception as e:
        return {
            'error': f'Error processing file: {str(e)}',
            'file_name': os.path.basename(file_path)
        }

def process_prd_format(file_path):
    """
    Process PRD format Excel file with merged cells for period names
    Row 1: Merged cells for period names (JANUARI 2024, FEBRUARI 2024, etc.)
    Row 2: Headers (NO, NIK, NAMA ANAK, TANGGAL LAHIR, JENIS KELAMIN, TGL UKUR, UMUR, etc.)
    Row 3+: Data
    """
    try:
        # Extract period names from merged cells
        period_names = extract_period_names_from_merged_cells(file_path)

        # Read data with openpyxl to properly handle the structure
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        # Determine the data structure
        max_row = ws.max_row
        max_col = ws.max_column

        # Get headers from row 2 (index 2 in openpyxl)
        headers = {}
        for col_idx in range(1, max_col + 1):
            cell_value = ws.cell(row=2, column=col_idx).value
            if cell_value:
                headers[col_idx - 1] = str(cell_value).strip()  # Convert to 0-based index

        # Find period columns based on headers (start after 6 identity columns)
        period_columns = []
        period_index = 0
        sub_columns = []

        for col_idx in range(7, max_col + 1):  # 1-based for openpyxl, start from column 7 (after 6 identity columns)
            header_str = headers.get(col_idx - 1, '')

            if header_str.upper() in ['TGL UKUR', 'UMUR', 'BERAT', 'TINGGI', 'CARA UKUR']:
                sub_columns.append((col_idx - 1, header_str))  # Convert to 0-based index

                # If we have all 5 sub-columns, create a period
                if len(sub_columns) == 5:
                    if period_index < len(period_names):
                        period_name = period_names[period_index]
                    else:
                        period_name = f'Periode {period_index + 1}'

                    period_columns.append({
                        'period_name': period_name,
                        'sub_columns': sub_columns.copy()
                    })
                    period_index += 1
                    sub_columns = []

        # Handle any remaining sub-columns
        if sub_columns:
            if period_index < len(period_names):
                period_name = period_names[period_index]
            else:
                period_name = f'Periode {period_index + 1}'

            period_columns.append({
                'period_name': period_name,
                'sub_columns': sub_columns
            })

        # Process data from row 3 onwards (index 3 in openpyxl)
        children = []
        for row_idx in range(3, max_row + 1):  # 1-based for openpyxl
            # Read the entire row
            row_data = []
            for col_idx in range(1, max_col + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                row_data.append(cell_value)

            # Convert to pandas Series for compatibility with extract_child_data
            import pandas as pd
            row_series = pd.Series(row_data)

            # Skip empty rows
            if row_series.isna().all():
                continue

            child_data = extract_child_data(row_series, period_columns, start_col=0)
            if child_data['nama_anak'] or child_data['nik']:
                # Apply WHO assessment and height validation
                apply_assessment_rules(child_data)
                children.append(child_data)

        wb.close()

        return {
            'file_name': os.path.basename(file_path),
            'format_type': 'PRD Format with Merged Cells',
            'total_children': len(children),
            'total_periods': len(period_columns),
            'periods': [p['period_name'] for p in period_columns],
            'children': children
        }

    except Exception as e:
        return {
            'error': f'Error processing PRD format: {str(e)}',
            'file_name': os.path.basename(file_path)
        }

def process_header_format(file_path):
    """
    Process header format Excel file (header TGL UKUR, UMUR, dll di baris 1)
    """
    try:
        df = pd.read_excel(file_path)
        data_rows = df.iloc[1:].copy()  # Skip header row
        data_rows = data_rows.reset_index(drop=True)

        # Find period columns from header
        header_row = df.iloc[0].fillna('')
        period_columns = []

        # Period names for sequential measurements
        period_names = ['Jan 2025', 'Feb 2025', 'Mar 2025', 'Apr 2025', 'May 2025', 'Jun 2025',
                       'Jul 2025', 'Aug 2025', 'Sep 2025', 'Oct 2025', 'Nov 2025', 'Dec 2025',
                       'Jan 2024', 'Feb 2024', 'Mar 2024', 'Apr 2024', 'May 2024', 'Jun 2024',
                       'Jul 2024', 'Aug 2024', 'Sep 2024', 'Oct 2024', 'Nov 2024', 'Dec 2024']

        sub_columns = []
        period_index = 0

        for i, header in enumerate(header_row):
            header_str = str(header).strip()

            if header_str in ['TGL UKUR', 'UMUR', 'BERAT', 'TINGGI', 'CARA UKUR']:
                sub_columns.append((i, header_str))

                # If we have all 5 sub-columns, create a period
                if len(sub_columns) == 5:
                    if period_index < len(period_names):
                        period_columns.append({
                            'period_name': period_names[period_index],
                            'sub_columns': sub_columns.copy()
                        })
                        period_index += 1
                    sub_columns = []

        # Handle any remaining sub-columns
        if sub_columns and period_index < len(period_names):
            period_columns.append({
                'period_name': period_names[period_index],
                'sub_columns': sub_columns
            })

        children = []
        for idx, row in data_rows.iterrows():
            if row.isna().all():
                continue

            child_data = extract_child_data(row, period_columns, start_col=0)
            if child_data['nama_anak'] or child_data['nik']:
                # Apply WHO assessment and height validation
                apply_assessment_rules(child_data)
                children.append(child_data)

        return {
            'file_name': os.path.basename(file_path),
            'format_type': 'Header Format',
            'total_children': len(children),
            'total_periods': len(period_columns),
            'periods': [p['period_name'] for p in period_columns],
            'children': children
        }

    except Exception as e:
        return {
            'error': f'Error processing header format: {str(e)}',
            'file_name': os.path.basename(file_path)
        }

def process_direct_data_format(file_path):
    """
    Process direct data format (data starts from first row)
    """
    try:
        df = pd.read_excel(file_path)

        children = []
        for idx, row in df.iterrows():
            if row.isna().all():
                continue

            # For direct format, assume structure: NO, NIK, NAMA, TGL LAHIR, JENIS KELAMIN, then measurements
            child_data = extract_child_data_direct_format(row)
            if child_data['nama_anak'] or child_data['nik']:
                # Apply WHO assessment and height validation
                apply_assessment_rules(child_data)
                children.append(child_data)

        return {
            'file_name': os.path.basename(file_path),
            'format_type': 'Direct Data Format',
            'total_children': len(children),
            'total_periods': 'Multiple (detected from data)',
            'periods': ['Jan 2025', 'Feb 2025', 'Mar 2025', 'Apr 2025', 'May 2025', 'Jun 2025', 'Jul 2025', 'Aug 2025', 'Sep 2025'],
            'children': children
        }

    except Exception as e:
        return {
            'error': f'Error processing direct data format: {str(e)}',
            'file_name': os.path.basename(file_path)
        }

def extract_child_data(row, period_columns, start_col=0):
    """
    Extract child data from a row given period columns configuration
    """
    child_data = {
        'no': None,
        'tempat': None,
        'nik': None,
        'nama_anak': None,
        'tanggal_lahir': None,
        'jenis_kelamin': None,
        'measurements': []
    }

    try:
        # Extract identity information (first 6 columns: NO, TEMPAT, NIK, NAMA ANAK, TANGGAL LAHIR, JENIS KELAMIN)
        if len(row) > start_col:
            child_data['no'] = int(row.iloc[start_col]) if not pd.isna(row.iloc[start_col]) else None
        if len(row) > start_col + 1:
            child_data['tempat'] = str(row.iloc[start_col + 1]).strip() if not pd.isna(row.iloc[start_col + 1]) else None
        if len(row) > start_col + 2:
            child_data['nik'] = str(row.iloc[start_col + 2]).strip() if not pd.isna(row.iloc[start_col + 2]) else None
        if len(row) > start_col + 3:
            child_data['nama_anak'] = str(row.iloc[start_col + 3]).strip() if not pd.isna(row.iloc[start_col + 3]) else None
        if len(row) > start_col + 4:
            if not pd.isna(row.iloc[start_col + 4]):
                if isinstance(row.iloc[start_col + 4], datetime):
                    child_data['tanggal_lahir'] = row.iloc[start_col + 4].strftime('%Y-%m-%d')
                else:
                    try:
                        date_obj = pd.to_datetime(row.iloc[start_col + 4])
                        child_data['tanggal_lahir'] = date_obj.strftime('%Y-%m-%d')
                    except:
                        child_data['tanggal_lahir'] = str(row.iloc[start_col + 4])
        if len(row) > start_col + 5:
            child_data['jenis_kelamin'] = str(row.iloc[start_col + 5]).strip().upper() if not pd.isna(row.iloc[start_col + 5]) else None

        # Process measurements
        for period in period_columns:
            measurement = {
                'periode': period['period_name'],
                'tgl_ukur': None,
                'umur_bulan': None,
                'berat_kg': None,
                'tinggi_cm': None,
                'cara_ukur': None
            }

            has_complete_data = False
            has_any_data = False

            for col_idx, sub_col_name in period['sub_columns']:
                if col_idx < len(row):
                    value = row.iloc[col_idx]
                    if not pd.isna(value):
                        has_any_data = True

                        if sub_col_name == 'TGL UKUR':
                            if isinstance(value, datetime):
                                measurement['tgl_ukur'] = value.strftime('%Y-%m-%d')
                            else:
                                try:
                                    date_obj = pd.to_datetime(value)
                                    measurement['tgl_ukur'] = date_obj.strftime('%Y-%m-%d')
                                except:
                                    measurement['tgl_ukur'] = str(value)
                        elif sub_col_name == 'UMUR':
                            try:
                                measurement['umur_bulan'] = int(float(value))
                            except (ValueError, TypeError):
                                measurement['umur_bulan'] = None
                        elif sub_col_name == 'BERAT':
                            try:
                                measurement['berat_kg'] = float(value)
                                has_complete_data = True  # Berat adalah data kunci
                            except (ValueError, TypeError):
                                measurement['berat_kg'] = None
                        elif sub_col_name == 'TINGGI':
                            try:
                                measurement['tinggi_cm'] = float(value)
                                has_complete_data = True  # Tinggi adalah data kunci
                            except (ValueError, TypeError):
                                measurement['tinggi_cm'] = None
                        elif sub_col_name == 'CARA UKUR':
                            measurement['cara_ukur'] = str(value).strip().upper()

            # Add status flags for UI display
            measurement['has_complete_data'] = has_complete_data
            measurement['is_incomplete'] = has_any_data and not has_complete_data

            # Always add measurement (complete or incomplete) for JSON output
            # UI will filter based on has_complete_data for count, but show all for transparency
            if has_any_data:
                child_data['measurements'].append(measurement)

    except Exception as e:
        print(f"Error extracting child data: {str(e)}")

    return child_data

def extract_child_data_direct_format(row):
    """
    Extract child data from direct format (no headers)
    """
    child_data = {
        'no': None,
        'nik': None,
        'nama_anak': None,
        'tanggal_lahir': None,
        'jenis_kelamin': None,
        'measurements': []
    }

    try:
        # Extract basic info (first 5 columns)
        if len(row) > 0:
            child_data['no'] = int(row.iloc[0]) if not pd.isna(row.iloc[0]) else None
        if len(row) > 1:
            child_data['nik'] = str(row.iloc[1]).strip() if not pd.isna(row.iloc[1]) else None
        if len(row) > 2:
            child_data['nama_anak'] = str(row.iloc[2]).strip() if not pd.isna(row.iloc[2]) else None
        if len(row) > 3:
            if isinstance(row.iloc[3], datetime):
                child_data['tanggal_lahir'] = row.iloc[3].strftime('%Y-%m-%d')
            else:
                try:
                    date_obj = pd.to_datetime(row.iloc[3])
                    child_data['tanggal_lahir'] = date_obj.strftime('%Y-%m-%d')
                except:
                    child_data['tanggal_lahir'] = str(row.iloc[3])
        if len(row) > 4:
            child_data['jenis_kelamin'] = str(row.iloc[4]).strip().upper() if not pd.isna(row.iloc[4]) else None

        # Extract measurements (assuming pattern: TGL UKUR, UMUR, BERAT, TINGGI, CARA UKUR for each period)
        period_names = ['Jan 2025', 'Feb 2025', 'Mar 2025', 'Apr 2025', 'May 2025', 'Jun 2025', 'Jul 2025', 'Aug 2025', 'Sep 2025']
        measurement_start_col = 5

        for i, period_name in enumerate(period_names):
            base_col = measurement_start_col + (i * 5)

            if base_col + 4 < len(row):
                measurement = {
                    'periode': period_name,
                    'tgl_ukur': None,
                    'umur_bulan': None,
                    'berat_kg': None,
                    'tinggi_cm': None,
                    'cara_ukur': None
                }

                # TGL UKUR
                if not pd.isna(row.iloc[base_col]):
                    if isinstance(row.iloc[base_col], datetime):
                        measurement['tgl_ukur'] = row.iloc[base_col].strftime('%Y-%m-%d')
                    else:
                        try:
                            date_obj = pd.to_datetime(row.iloc[base_col])
                            measurement['tgl_ukur'] = date_obj.strftime('%Y-%m-%d')
                        except:
                            measurement['tgl_ukur'] = str(row.iloc[base_col])

                # UMUR
                if not pd.isna(row.iloc[base_col + 1]):
                    try:
                        measurement['umur_bulan'] = int(float(row.iloc[base_col + 1]))
                    except (ValueError, TypeError):
                        measurement['umur_bulan'] = None

                # BERAT
                if not pd.isna(row.iloc[base_col + 2]):
                    try:
                        measurement['berat_kg'] = float(row.iloc[base_col + 2])
                    except (ValueError, TypeError):
                        measurement['berat_kg'] = None

                # TINGGI
                if not pd.isna(row.iloc[base_col + 3]):
                    try:
                        measurement['tinggi_cm'] = float(row.iloc[base_col + 3])
                    except (ValueError, TypeError):
                        measurement['tinggi_cm'] = None

                # CARA UKUR
                if not pd.isna(row.iloc[base_col + 4]):
                    measurement['cara_ukur'] = str(row.iloc[base_col + 4]).strip().upper()

                # Check if measurement has complete data (berat or tinggi)
                has_complete_data = (measurement['berat_kg'] is not None or measurement['tinggi_cm'] is not None)
                has_any_data = (measurement['tgl_ukur'] is not None or measurement['umur_bulan'] is not None or
                               measurement['berat_kg'] is not None or measurement['tinggi_cm'] is not None or
                               measurement['cara_ukur'] is not None)

                # Add status flags for UI display
                measurement['has_complete_data'] = has_complete_data
                measurement['is_incomplete'] = has_any_data and not has_complete_data

                # Always add measurement (complete or incomplete) for JSON output
                if has_any_data:
                    child_data['measurements'].append(measurement)

    except Exception as e:
        print(f"Error extracting child data direct format: {str(e)}")

    return child_data

def save_json_to_file(data, output_path):
    """
    Save JSON data to file
    """
    try:
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        return True
    except Exception as e:
        print(f'Error saving JSON: {str(e)}')
        return False

def detect_excel_format(file_path):
    """
    Detect the format of Excel file
    Returns: format_type, description
    """
    try:
        # First check for merged cells PRD format
        try:
            period_names = extract_period_names_from_merged_cells(file_path)
            if period_names and len(period_names) > 0:
                # Use openpyxl to check the actual structure
                wb = openpyxl.load_workbook(file_path)
                ws = wb.active

                # Check row 2 (index 2 in openpyxl) for identity headers
                identity_headers = []
                for col_idx in range(1, 7):  # Check first 6 columns (including TEMPAT)
                    cell_value = ws.cell(row=2, column=col_idx).value  # Row 2 = index 2
                    if cell_value:
                        identity_headers.append(str(cell_value).strip().upper())

                wb.close()

                required_identity = ['NO', 'NIK', 'NAMA ANAK', 'TANGGAL LAHIR', 'JENIS KELAMIN', 'TEMPAT']
                identity_count = sum(1 for col in required_identity if col in identity_headers)

                if identity_count >= 3:
                    return 'prd_format', f'Format PRD dengan merged cells ({len(period_names)} periode terdeteksi)'
        except Exception as e:
            print(f"Debug: Error checking merged cells: {e}")
            pass  # Continue with other detection methods

        df = pd.read_excel(file_path)

        # Check if first row looks like headers (contains strings like 'TGL UKUR', 'UMUR', etc.)
        first_row = df.iloc[0].fillna('').astype(str)
        header_keywords = ['TGL UKUR', 'UMUR', 'BERAT', 'TINGGI', 'CARA UKUR']

        header_count = sum(1 for keyword in header_keywords if keyword in first_row.values)

        if header_count >= 3:
            return 'header_format', 'Format dengan header di baris pertama'

        # Check if first row contains data (numbers, names, dates)
        first_col_values = first_row.iloc[:5].tolist()
        has_data = False

        for val in first_col_values:
            if (val and
                (val.replace('.', '').replace('-', '').isdigit() or  # Numbers
                 len(val) > 3 or  # Names (length > 3)
                 any(char in val for char in ['-', '/']))):  # Dates
                has_data = True
                break

        if has_data:
            return 'direct_data', 'Format data langsung tanpa header'

        # Check traditional PRD format (headers in first row, sub-headers in second row)
        if len(df) >= 2:
            main_header = df.iloc[0].fillna('').astype(str)
            required_identity = ['NO', 'NIK', 'NAMA ANAK', 'TANGGAL LAHIR', 'JENIS KELAMIN']

            identity_count = sum(1 for col in required_identity if col in main_header.values)

            if identity_count >= 3:  # At least some identity columns found
                return 'prd_format', 'Format PRD (header di baris 1, sub-header di baris 2)'

        return 'unknown', 'Format tidak dikenali'

    except Exception as e:
        return 'error', f'Error detecting format: {str(e)}'

def validate_excel_format(file_path):
    """
    Validate if Excel file follows a supported format
    """
    format_type, message = detect_excel_format(file_path)

    if format_type == 'error':
        return False, message

    if format_type == 'unknown':
        return False, "Format Excel tidak dikenali. Harap gunakan format PRD, header, atau data langsung."

    return True, f"Format terdeteksi: {message}"

def validate_template_compliance(file_path):
    """
    Flexible template validation - accepts both PRD and current Data Test.xlsx format
    Returns: (is_valid, validation_result)
    """
    try:
        df = pd.read_excel(file_path)

        # Check minimum data requirements
        if df.empty:
            return False, {
                'valid': False,
                'errors': ['File Excel kosong'],
                'warnings': [],
                'format_detected': 'unknown',
                'needs_template_download': True
            }

        validation_result = {
            'valid': True,
            'errors': [],
            'warnings': [],
            'format_detected': None,
            'needs_template_download': False
        }

        # Detect format
        format_type, format_description = detect_excel_format(file_path)
        validation_result['format_detected'] = format_type

        # Accept both PRD format and current Data Test.xlsx format (which is detected as header_format)
        if format_type not in ['prd_format', 'header_format']:
            validation_result['valid'] = False
            validation_result['errors'] = [
                'File tidak sesuai template yang disepakati. Harap gunakan format template yang benar.',
                'Download template reference "Data Test.xlsx" untuk format yang benar.'
            ]
            validation_result['needs_template_download'] = True
            return False, validation_result

        # For PRD format, check identity columns in first row
        if format_type == 'prd_format':
            main_header = df.iloc[0].fillna('').astype(str)
            required_identity = ['NO', 'NIK', 'NAMA ANAK', 'TANGGAL LAHIR', 'JENIS KELAMIN']

            # Check if all required identity columns are present
            missing_identity = []
            for col in required_identity:
                if col not in main_header.values:
                    missing_identity.append(col)

            if missing_identity:
                validation_result['valid'] = False
                validation_result['errors'] = [
                    f'Template tidak lengkap. Kolom wajib yang hilang: {", ".join(missing_identity)}',
                    'Harap gunakan template "Data Test.xlsx" sebagai referensi.'
                ]
                validation_result['needs_template_download'] = True
                return False, validation_result

        # For header format (current Data Test.xlsx), check for proper structure
        if format_type == 'header_format':
            # Check if we have proper header structure with TGL UKUR, UMUR, etc.
            header_row = df.iloc[0].fillna('').astype(str)
            required_headers = ['TGL UKUR', 'UMUR', 'BERAT', 'TINGGI', 'CARA UKUR']
            found_headers = [h for h in required_headers if h in header_row.values]

            if len(found_headers) < 3:
                validation_result['warnings'].append(f'Hanya ditemukan {len(found_headers)} dari 5 header pengukuran yang diharapkan')

        # Check if there's data (at least 2 rows)
        if len(df) < 2:
            validation_result['valid'] = False
            validation_result['errors'] = [
                'File harus memiliki minimal 1 baris data anak',
                'Harap gunakan template "Data Test.xlsx" sebagai referensi.'
            ]
            validation_result['needs_template_download'] = True
            return False, validation_result

        return validation_result['valid'], validation_result

    except Exception as e:
        return False, {
            'valid': False,
            'errors': [f'Error reading file: {str(e)}'],
            'warnings': [],
            'format_detected': 'error',
            'needs_template_download': True
        }


if __name__ == '__main__':
    # Example usage
    test_files = [
        'data_template/Data Test.xlsx',
        'data_template/Data Test Header.xlsx'
    ]

    for test_file in test_files:
        if os.path.exists(test_file):
            print(f"\n=== Processing {test_file} ===")

            # Validate format first
            is_valid, message = validate_excel_format(test_file)
            print(f"Validation: {message}")

            if is_valid:
                result = process_excel_to_json(test_file)

                if 'error' in result:
                    print(f"Error: {result['error']}")
                else:
                    print(f"Success: {result['total_children']} children processed")
                    print(f"Periods found: {', '.join(result['periods'])}")

                    # Show first child as example
                    if result['children']:
                        print("\nExample child data:")
                        print(json.dumps(result['children'][0], indent=2))

                    # Save to JSON file
                    output_file = f"output_{os.path.basename(test_file).replace('.xlsx', '.json').replace('.xls', '.json')}"
                    if save_json_to_file(result, output_file):
                        print(f"\nJSON saved to {output_file}")
            else:
                print(f"Skipping invalid file: {message}")
        else:
            print(f"File {test_file} not found")